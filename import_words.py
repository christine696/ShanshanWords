# -*- coding: utf-8 -*-
"""词库导入脚本 - 导入 GRE 词汇和 IELTS 场景词汇"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import csv
import re
import openpyxl
import pdfplumber

# ============================================================
# 1. 从 Excel 读取 GRE 词汇
# ============================================================
def import_gre_words():
    gre_words = []
    excel_path = 'd:/Desktop/gre核心词data.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 使用 "词表" 工作表
    if '词表' in wb.sheetnames:
        ws = wb['词表']
    else:
        ws = wb.active

    # 跳过前两行标题
    for row in ws.iter_rows(min_row=3, values_only=True):
        word = row[0]
        if not word or not isinstance(word, str):
            continue
        word = word.strip().lower()
        if not word:
            continue

        pos = str(row[1]).strip() if row[1] else ''
        meaning = str(row[2]).strip() if row[2] else ''
        difficulty_raw = row[3]

        # difficulty 映射
        try:
            difficulty = int(difficulty_raw) if difficulty_raw else 1
        except (ValueError, TypeError):
            difficulty = 1

        gre_words.append({
            'word': word,
            'category': 'GRE',
            'subcategory': 'GRE核心词',
            'phonetic': '',
            'part_of_speech': pos,
            'meaning_cn': meaning,
            'example_en': '',
            'example_cn': '',
            'source': 'gre核心词data.xlsx',
            'difficulty': difficulty
        })

    wb.close()
    return gre_words

# ============================================================
# 2. 从 PDF 读取 IELTS 场景词汇
# ============================================================
def import_ielts_scene_words():
    """从 PDF 提取听力词汇篇的场景词汇"""
    scene_words = []
    pdf_path = 'd:/Desktop/雅思场景超核心词汇4000.pdf'

    # 场景定义：(场景名, 起始页, 结束页)
    # 基于目录和之前检查的结果
    scenes = [
        ('租房场景', 7, 16),
        ('旅游场景', 10, 19),
        ('餐饮场景', 13, 22),
        ('节日及活动场景', 16, 24),
        ('工作场景', 18, 26),
        ('银行场景', 20, 28),
        ('图书馆场景', 22, 30),
        ('选课场景', 24, 35),
        ('作业讨论场景', 29, 40),
    ]

    def parse_word_line(line):
        """解析单词行，如 'motel/məʊˈtel/n. 汽车旅馆' 或 'air-conditionern. 空调'"""
        line = line.strip()
        if not line or line.startswith('一、') or line.startswith('二、') or line.startswith('三、') or \
           line.startswith('四、') or line.startswith('五、') or line.startswith('六、'):
            return None

        # 移除开头的圆点符号
        line = re.sub(r'^[●○•\*\s]+', '', line)
        if not line:
            return None

        # 提取音标 (在 / / 或 [ ] 之间)
        phonetic = ''
        ph_match = re.search(r'/([^/]+)/', line)
        if ph_match:
            phonetic = ph_match.group(1)
            line = line.replace(ph_match.group(0), '')

        ph_match2 = re.search(r'\[([^\]]+)\]', line)
        if ph_match2 and not phonetic:
            phonetic = ph_match2.group(1)
            line = line.replace(ph_match2.group(0), '')

        # 提取词性
        pos = ''
        pos_match = re.search(r'\b(n\.?|v\.?|adj\.?|adv\.?|conj\.?|prep\.?|pron\.?)\b', line)
        if pos_match:
            pos = pos_match.group(1).rstrip('.')

        # 剩余部分是单词和中文释义
        # 格式可能是: 单词/词/音标/n. 释义  或  单词/n. 释义
        # 提取英文单词（第一串不含中文的字符串）
        en_part = re.split(r'[/一-鿿]', line)[0].strip()
        # 去掉可能的数字开头
        en_part = re.sub(r'^\d+\.\s*', '', en_part)

        # 中文释义
        cn_match = re.search(r'[一-鿿][^\n]*', line)
        meaning_cn = cn_match.group(0).strip() if cn_match else ''

        if not en_part or not meaning_cn:
            return None

        return {
            'word': en_part.strip(),
            'phonetic': phonetic.strip(),
            'part_of_speech': pos.strip(),
            'meaning_cn': meaning_cn.strip()
        }

    with pdfplumber.open(pdf_path) as pdf:
        for scene_name, start_page, end_page in scenes:
            for page_num in range(start_page, min(end_page, len(pdf.pages))):
                text = pdf.pages[page_num].extract_text()
                if not text:
                    continue

                # 检查这一页是否还在当前场景
                # 用目录页判断是否超出场景范围（简单策略：检查页码变化）
                if scene_name not in text:
                    # 可能已经进入下一个场景了
                    # 检查是否进入了下一个场景
                    found_next = False
                    for next_scene, ns, ne in scenes:
                        if next_scene != scene_name and next_scene in text:
                            found_next = True
                            break
                    if found_next:
                        break

                lines = text.split('\n')
                for line in lines:
                    parsed = parse_word_line(line)
                    if parsed:
                        scene_words.append({
                            'word': parsed['word'],
                            'category': 'IELTS_SCENE',
                            'subcategory': scene_name,
                            'phonetic': parsed['phonetic'],
                            'part_of_speech': parsed['part_of_speech'],
                            'meaning_cn': parsed['meaning_cn'],
                            'example_en': '',
                            'example_cn': '',
                            'source': '雅思场景超核心词汇4000.pdf',
                            'difficulty': 1
                        })

    return scene_words

# ============================================================
# 3. 读取原有 CSV
# ============================================================
def import_existing_csv():
    existing_words = []
    csv_path = 'data/words.csv'
    if not os.path.exists(csv_path):
        return existing_words

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            word = row.get('word', '').strip()
            category = row.get('category', '').strip()
            if not word or not category:
                continue
            existing_words.append({
                'word': word,
                'category': category,
                'subcategory': row.get('subcategory', 'IELTS核心词' if category == 'IELTS' else 'GRE核心词'),
                'phonetic': row.get('phonetic', ''),
                'part_of_speech': row.get('part_of_speech', ''),
                'meaning_cn': row.get('meaning_cn', ''),
                'example_en': row.get('example_en', ''),
                'example_cn': row.get('example_cn', ''),
                'source': row.get('source', ''),
                'difficulty': int(row.get('difficulty', 1) or 1)
            })
    return existing_words

# ============================================================
# 4. 合并去重
# ============================================================
def merge_and_deduplicate(all_words):
    """按 word + category + subcategory 去重，保留信息更完整的"""
    seen = {}
    for w in all_words:
        key = (w['word'].lower(), w['category'], w.get('subcategory', ''))
        if key not in seen:
            seen[key] = w
        else:
            # 保留信息更完整的
            existing = seen[key]
            for field in ['phonetic', 'part_of_speech', 'meaning_cn', 'example_en', 'example_cn']:
                if not existing.get(field) and w.get(field):
                    existing[field] = w[field]
            # difficulty 保留较小值（较难的优先）
            if w.get('difficulty', 1) > existing.get('difficulty', 1):
                existing['difficulty'] = w['difficulty']

    return list(seen.values())

# ============================================================
# 5. 写入 CSV
# ============================================================
def write_master_csv(words):
    fieldnames = ['word', 'category', 'subcategory', 'phonetic', 'part_of_speech',
                  'meaning_cn', 'example_en', 'example_cn', 'source', 'difficulty']

    output_path = 'data/words_master.csv'
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        # 按 category, subcategory, word 排序
        words_sorted = sorted(words, key=lambda x: (x['category'], x.get('subcategory', ''), x['word'].lower()))
        writer.writerows(words_sorted)

    return output_path

# ============================================================
# 主程序
# ============================================================
if __name__ == '__main__':
    print('开始导入词库...')

    # 读取原有词库
    print('1. 读取原有词库...')
    existing = import_existing_csv()
    print(f'   原有词汇: {len(existing)} 个')

    # 导入 GRE
    print('2. 导入 GRE 词汇...')
    gre_words = import_gre_words()
    print(f'   GRE 词汇: {len(gre_words)} 个')

    # 导入 IELTS 场景
    print('3. 导入 IELTS 场景词汇...')
    ielts_scene = import_ielts_scene_words()
    print(f'   IELTS 场景词汇: {len(ielts_scene)} 个')

    # 合并
    print('4. 合并去重...')
    all_words = existing + gre_words + ielts_scene
    print(f'   合并前总数: {len(all_words)}')
    merged = merge_and_deduplicate(all_words)
    print(f'   去重后总数: {len(merged)}')

    # 写入 CSV
    print('5. 写入词库文件...')
    output_path = write_master_csv(merged)
    print(f'   已保存到: {output_path}')

    # 统计
    from collections import Counter
    counts = Counter(w['category'] for w in merged)
    print('\n=== 导入统计 ===')
    for cat, cnt in sorted(counts.items()):
        print(f'  {cat}: {cnt} 个')

    # 按子分类统计
    subcat_counts = Counter((w['category'], w.get('subcategory', '')) for w in merged)
    print('\n=== 子分类统计 ===')
    for (cat, sub), cnt in sorted(subcat_counts.items()):
        print(f'  {cat} / {sub}: {cnt} 个')
